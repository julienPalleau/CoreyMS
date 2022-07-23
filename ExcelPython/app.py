import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def list_files():
    path = Path()
    for file in path.glob('*.*'):  # liste tous les fichiers
        print(file)


def pickup_your_file():
    choosen_file = input('choisissez votre fichier :')
    return choosen_file


def process_workbook(choosen_file):
    wb = xl.load_workbook(choosen_file)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):  # we start from 2 as we don't want the first line (line with titles)
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('transactions2.xlsx')


list_files()
process_workbook(pickup_your_file())
