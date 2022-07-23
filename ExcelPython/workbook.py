from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")  # insert at the end (default)
ws2 = wb.create_sheet("Mysheet", 0)  # insert at first position
ws3 = wb.create_sheet("Mysheet", -1)  # insert at the last position
ws.title = "Julien's tests in python"
ws3 = wb["Julien's tests in python"]
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)


c = ws['A1']
d = ws.cell(row=4, column=2, value=10)
print(c.value)
print(d.value)

ws['A4'] = 4
ws.cell(row=1, column=2, value=15)

c.value = 'hello, world'
d.value = 3.14

print("Printing cells")
for x in range(1, 10):
    for y in range(1, 10):
        ws.cell(row=x, column=y)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

for row in ws.values:
    for value in row:
        print(value)



wb.save('test.xlsx')

